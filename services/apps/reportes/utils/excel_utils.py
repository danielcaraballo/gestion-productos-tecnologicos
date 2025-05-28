from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def generate_excel_report(data):
    """Genera reporte Excel con datos proporcionados"""
    wb = Workbook()

    # Creación de hojas
    _create_summary_sheet(wb, data)
    _create_detail_sheet(wb, data)
    _create_status_sheets(wb, data)

    # Ajustes finales
    _apply_final_adjustments(wb)

    return wb


def _create_summary_sheet(wb, data):
    """Configura la hoja de resumen principal"""
    ws = wb.active
    ws.title = "Resumen"

    # Definir el estilo de título
    estilo_titulo = Font(name='Calibri', bold=True, size=12)
    alineacion_centrada = Alignment(horizontal='center')

    # Título principal
    ws.append(["PORTAFOLIO DE PRODUCTOS TECNOLÓGICOS"])
    ws.merge_cells('A1:B1')
    ws['A1'].font = estilo_titulo
    ws['A1'].alignment = alineacion_centrada

    ws.append(["Oficina de Tecnologías de la Información y la Comunicación"])
    ws.merge_cells('A2:B2')
    ws.append([])

    # Estadísticas básicas
    total_cell = f"Total de Productos: {data['stats']['total_productos']}"
    ws.append([total_cell])
    ws.merge_cells(f'A{ws.max_row}:B{ws.max_row}')
    ws[f'A{ws.max_row}'].font = estilo_titulo
    ws[f'A{ws.max_row}'].alignment = alineacion_centrada
    ws.append([])

    # Secciones de distribución
    distributions = [
        ('por_estatus', 'estatus__nombre', 'Distribución por Estatus'),
        ('por_categoria', 'categoria__nombre', 'Distribución por Categoría'),
        ('por_dependencia_usuaria', 'dependencia_usuaria__nombre',
         'Distribución por Dependencia')
    ]

    for key, field, title in distributions:
        ws.append([title])
        ws.merge_cells(f'A{ws.max_row}:B{ws.max_row}')
        ws[f'A{ws.max_row}'].font = estilo_titulo
        ws[f'A{ws.max_row}'].alignment = alineacion_centrada

        if key in data['stats']:
            for item in data['stats'][key]:
                name = item[field] or "Sin especificar"
                ws.append([name, item['total']])

        ws.append([])

    # Pie de página
    ws.append(
        ["Reporte generado por la Aplicación de Gestión de Productos Tecnológicos"])
    ws.merge_cells(f'A{ws.max_row}:B{ws.max_row}')

    # Ajuste de columnas (solo para hoja Resumen)
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 5


def _create_detail_sheet(wb, data):
    """Crea hoja con detalle completo de productos"""
    ws = wb.create_sheet("Detalle")

    # Definir el estilo de título
    estilo_titulo = Font(name='Calibri', bold=True, size=12)
    alineacion_centrada = Alignment(horizontal='center')
    estilo_encabezado = Font(name='Calibri', bold=True, size=11)

    # Título de la hoja y encabezados SIN fila vacía intermedia
    ws.append(["DETALLE DEL PORTAFOLIO DE PRODUCTOS TECNOLÓGICOS"])
    ws.merge_cells('A1:G1')
    ws['A1'].font = estilo_titulo
    ws['A1'].alignment = alineacion_centrada

    # Encabezados de tabla (fila inmediatamente después del título)
    headers = [
        "Nombre", "Descripción", "Categoría",
        "Dependencia", "Subdependencia",
        "Componentes", "URLs"
    ]
    ws.append(headers)

    # Aplicar estilo a los encabezados
    for col in range(1, 8):
        celda = ws.cell(row=2, column=col)  # Fila 2 (antes era 3)
        celda.font = estilo_encabezado
        celda.alignment = alineacion_centrada

    # Datos de productos
    for prod in data['productos']:
        components = "\n".join(f"{c.nombre} ({c.tipo.nombre})" for c in prod.componentes.all(
        )) if prod.componentes.exists() else "N/A"
        urls = "\n".join(c.direccion_url for c in prod.componentes.all(
        )) if prod.componentes.exists() else "N/A"

        ws.append([
            prod.nombre,
            prod.descripcion,
            prod.categoria.nombre if prod.categoria else "No especificado",
            prod.dependencia_usuaria.nombre if prod.dependencia_usuaria else "No especificado",
            prod.subdependencia_usuaria.nombre if prod.subdependencia_usuaria else "No especificado",
            components,
            urls
        ])


def _create_status_sheets(wb, data):
    """Crea hojas separadas por estatus de producto"""
    # Definir el estilo de título
    estilo_titulo = Font(name='Calibri', bold=True, size=12)
    alineacion_centrada = Alignment(horizontal='center')
    estilo_encabezado = Font(name='Calibri', bold=True, size=11)

    # Agrupación por estatus
    status_groups = {}
    for prod in data['productos']:
        status = prod.estatus.nombre if prod.estatus else "Sin Estatus"
        status_groups.setdefault(status, []).append(prod)

    # Crear hoja por cada grupo
    for status, products in status_groups.items():
        ws = wb.create_sheet(_clean_sheet_name(status))

        # Título de la hoja y encabezados SIN fila vacía intermedia
        ws.append([f"PRODUCTOS - {status.upper()}"])
        ws.merge_cells('A1:G1')
        ws['A1'].font = estilo_titulo
        ws['A1'].alignment = alineacion_centrada

        # Encabezados de tabla (fila inmediatamente después del título)
        headers = [
            "Nombre", "Descripción", "Categoría",
            "Dependencia", "Subdependencia",
            "Componentes", "URLs"
        ]
        ws.append(headers)

        # Aplicar estilo a los encabezados
        for col in range(1, 8):
            celda = ws.cell(row=2, column=col)  # Fila 2 (antes era 3)
            celda.font = estilo_encabezado
            celda.alignment = alineacion_centrada

        # Agregar productos
        for prod in products:
            components = "\n".join(f"{c.nombre} ({c.tipo.nombre})" for c in prod.componentes.all(
            )) if prod.componentes.exists() else "N/A"
            urls = "\n".join(c.direccion_url for c in prod.componentes.all(
            )) if prod.componentes.exists() else "N/A"

            ws.append([
                prod.nombre,
                prod.descripcion,
                prod.categoria.nombre if prod.categoria else "No especificado",
                prod.dependencia_usuaria.nombre if prod.dependencia_usuaria else "No especificado",
                prod.subdependencia_usuaria.nombre if prod.subdependencia_usuaria else "No especificado",
                components,
                urls
            ])


def _clean_sheet_name(name):
    """Limpia nombre para hoja Excel (max 31 chars, sin caracteres inválidos)"""
    invalid_chars = [':', '\\', '?', '/']
    for char in invalid_chars:
        name = name.replace(char, '')
    return name[:31]


def _apply_final_adjustments(wb):
    """Aplica ajustes de formato a todas las hojas excepto Resumen"""
    for sheet in wb:
        if sheet.title != "Resumen":  # Solo aplicar a hojas que no son Resumen
            # Autoajuste de columnas mejorado
            for col in sheet.columns:
                col_letter = get_column_letter(col[0].column)
                max_len = 0
                for cell in col:
                    try:
                        if cell.value:
                            # Considerar saltos de línea para celdas con texto largo
                            if isinstance(cell.value, str):
                                lines = cell.value.split('\n')
                                line_len = max(len(line) for line in lines)
                                max_len = max(max_len, line_len)
                            else:
                                max_len = max(max_len, len(str(cell.value)))
                    except:
                        continue

                # Ajustar ancho con márgenes y límites
                if max_len > 0:
                    adjusted_width = (max_len + 2) * 1.1
                    sheet.column_dimensions[col_letter].width = min(
                        max(adjusted_width, 10), 50)

            # Congelar encabezados (ahora en fila 2)
            sheet.freeze_panes = "A3"  # Congela hasta la fila 2 (encabezados)
